import streamlit as st
from datetime import date, timedelta
import db_utils
import pandas as pd
import plotly.express as px  # 替换matplotlib为Plotly
from io import BytesIO
import jwt
import time
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# 初始化数据库
db_utils.init_db()

# 页面配置
st.set_page_config(page_title="工作记录管理系统", layout="wide")

# 获取数据库会话
def get_db():
    return next(db_utils.get_db_session())

# 检查JWT并自动续期
def check_auth():
    # 先检查URL参数中的token
    if 'token' in st.query_params and 'jwt_token' not in st.session_state:
        st.session_state.jwt_token = st.query_params["token"]
    
    if 'jwt_token' not in st.session_state:
        return False
    
    # 验证当前Token
    username = db_utils.verify_jwt_token(st.session_state.jwt_token)
    
    # Token无效或过期
    if not username:
        # 尝试使用URL参数中的token
        if 'token' in st.query_params:
            username = db_utils.verify_jwt_token(st.query_params["token"])
            if username:
                st.session_state.jwt_token = st.query_params["token"]
                st.session_state.username = username
                return True
        return False
    
    # 检查Token是否需要续期（剩余时间小于5分钟）
    try:
        payload = jwt.decode(
            st.session_state.jwt_token, 
            db_utils.SECRET_KEY, 
            algorithms=['HS256'],
            options={"verify_exp": False}  # 不验证过期时间
        )
        exp_time = payload['exp']
        if exp_time - time.time() < 300:  # 剩余时间小于5分钟
            # 生成新Token
            st.session_state.jwt_token = db_utils.generate_jwt_token(username)
            st.query_params["token"] = st.session_state.jwt_token
    except:
        pass
    
    # 确保username也在session state中
    if 'username' not in st.session_state:
        st.session_state.username = username
        
    # 检查前一天未完成的工作
    if username:
        yesterday = date.today() - timedelta(days=1)
        db = next(db_utils.get_db_session())
        uncompleted = db_utils.get_uncompleted_records(db, yesterday)
        if uncompleted:
            st.session_state.pending_records = uncompleted
            st.session_state.show_pending_records = True
    
    return True

# 登录/注册页面
if not check_auth():
    # 添加新的标题显示方式
    st.markdown("## 📊 工作记录管理系统 - 登录")
    
    tab_login, tab_register, tab_forgot = st.tabs(["🔐 登录", "📝 注册", "🔑 找回密码"])
    
    with tab_login:
        with st.form("login_form"):
            username = st.text_input("用户名")
            password = st.text_input("密码", type="password")
            remember = st.checkbox("记住我")
            
            if st.form_submit_button("登录"):
                db = get_db()
                user = db_utils.verify_user(db, username, password)
                if user:
                    # 生成JWT Token
                    token = db_utils.generate_jwt_token(username)
                    st.session_state.jwt_token = token
                    st.session_state.username = username
                    st.query_params["token"] = token
                    
                    # 修改: 使用新方法获取所有未完成记录（不限定日期）
                    uncompleted = db_utils.get_uncompleted_records(db)  # 删除日期参数
                    if uncompleted:
                        st.session_state.pending_records = uncompleted
                        st.session_state.show_pending_records = True
                        
                        # 添加调试信息
                        st.toast("⚠️ 检测到未完成工作，请及时处理！", icon='⚠️')
                    
                    st.success("登录成功！")
                    st.rerun()
                else:
                    st.error("用户名或密码错误")
    
    with tab_register:
        with st.form("register_form"):
            new_username = st.text_input("用户名")
            new_password = st.text_input("密码", type="password")
            confirm_password = st.text_input("确认密码", type="password")
            
            if st.form_submit_button("注册"):
                if new_password != confirm_password:
                    st.error("两次输入的密码不一致")
                else:
                    db = get_db()
                    user = db_utils.create_user(db, new_username, new_password)
                    if user:
                        st.success("注册成功！请登录")
                    else:
                        st.error("用户名已存在")
    
    with tab_forgot:
        with st.form("forgot_form"):
            forgot_username = st.text_input("用户名")
            new_password = st.text_input("新密码", type="password")
            confirm_password = st.text_input("确认新密码", type="password")
            
            if st.form_submit_button("重置密码"):
                if new_password != confirm_password:
                    st.error("两次输入的密码不一致")
                else:
                    db = get_db()
                    if db_utils.update_password(db, forgot_username, new_password):
                        st.success("密码已重置，请使用新密码登录")
                    else:
                        st.error("用户名不存在")
    
    st.stop()

# 主界面重构
st.title(f"工作记录管理系统 - 欢迎 {st.session_state.username}")

# 添加全局样式优化
st.markdown("""
<style>
    /* 卡片式容器样式 */
    .card {
        padding: 2rem;
        border-radius: 1.5rem;
        box-shadow: 0 6px 12px rgba(0, 0, 0, 0.15);
        background-color: #ffffff;
        margin: 1.5rem 0;
        transition: transform 0.3s ease, box-shadow 0.3s ease;
    }
    .card:hover {
        transform: translateY(-8px);
        box-shadow: 0 10px 20px rgba(0, 0, 0, 0.2);
    }
    
    /* 增强型按钮样式 */
    .stButton button {
        border-radius: 1rem;
        padding: 0.8rem 1.5rem;
        font-weight: 700;
        text-transform: uppercase;
        letter-spacing: 1px;
        background: linear-gradient(45deg, #3b82f6, #6366f1, #8b5cf6);
        color: white;
        border: none;
        transition: all 0.4s ease;
        box-shadow: 0 4px 14px rgba(59, 130, 246, 0.4);
    }
    .stButton button:hover {
        transform: scale(1.03);
        box-shadow: 0 6px 18px rgba(59, 130, 246, 0.6);
    }
    
    /* 数据表格样式优化 */
    .stDataFrame {
        border-radius: 1rem;
        overflow: hidden;
        box-shadow: 0 4px 8px rgba(0, 0, 0, 0.08);
    }
    .stDataFrame thead th {
        background-color: #bfdbfe !important;
        color: #1e40af !important;
        font-weight: 600;
    }
    .stDataFrame tbody tr:nth-child(even) {
        background-color: #efefef !important;
    }
    
    /* 日期选择器样式优化 */
    .stDateInput input {
        border-radius: 1rem;
        padding: 0.6rem 1.2rem;
        border: 2px solid #bfdbfe;
        font-size: 1rem;
        transition: all 0.3s ease;
    }
    .stDateInput input:focus {
        border-color: #3b82f6;
        box-shadow: 0 0 0 2px rgba(59, 130, 246, 0.2);
    }
    
    /* 分页控件样式 */
    .stNumberInput input {
        border-radius: 0.75rem;
        padding: 0.4rem 0.8rem;
        border: 1px solid #d1d5db;
        font-weight: 500;
    }
    
    /* 提醒卡片样式优化 */
    .reminder-card {
        padding: 1rem;
        border-radius: 1rem;
        margin: 0.5rem 0;
        background: linear-gradient(to right, #fef3c7, #fee3a1);
        border-left: 5px solid #f59e0b;
        box-shadow: 0 2px 6px rgba(0, 0, 0, 0.1);
    }
    .reminder-card:hover {
        transform: translateX(5px);
    }
    
    /* 图标动画效果 */
    .icon-animation {
        animation: pulse 2s infinite;
    }
    @keyframes pulse {
        0% { transform: scale(1); }
        50% { transform: scale(1.1); }
        100% { transform: scale(1); }
    }
</style>
""", unsafe_allow_html=True)

# 退出登录按钮（添加图标和样式优化）
st.markdown("""
<style>
    .logout-button {
        background: linear-gradient(45deg, #ff6b6b, #ff8e53);
        color: white;
        border-radius: 10px;
        padding: 8px 15px;
        font-weight: bold;
    }
    .logout-button:hover {
        transform: scale(1.05);
    }
</style>
""", unsafe_allow_html=True)

if st.button("🚪 退出登录", key="logout_button", help="点击退出系统", use_container_width=True):
    st.session_state.pop('jwt_token', None)
    st.session_state.pop('username', None)
    st.query_params.clear()  # 修改为使用query_params.clear()
    st.rerun()

# 新增系统管理页面
tab_main, tab_admin = st.tabs(["📊 工作记录", "⚙️ 系统管理"])

with tab_admin:
    # 用户管理
    with st.expander("用户管理"):
        st.subheader("用户列表")
        db = get_db()
        users = db_utils.get_all_users(db)
        
        if users:
            # 显示用户表格
            user_df = pd.DataFrame([{
                "ID": u.id,
                "用户名": u.username,
                "上次登录": u.last_login
            } for u in users])
            st.dataframe(user_df)
        else:
            st.warning("暂无用户")
        
        # 添加新用户
        st.subheader("添加用户")
        with st.form("add_user_form"):
            new_username = st.text_input("用户名")
            new_password = st.text_input("密码", type="password")
            confirm_password = st.text_input("确认密码", type="password")
            
            if st.form_submit_button("添加"):
                if new_password != confirm_password:
                    st.error("两次输入的密码不一致")
                else:
                    db = get_db()
                    user = db_utils.create_user(db, new_username, new_password)
                    if user:
                        st.success(f"用户 {new_username} 添加成功")
                        st.rerun()
                    else:
                        st.error("用户名已存在")
        
        # 修改密码
        st.subheader("修改密码")
        if users:
            usernames = [u.username for u in users]
            selected_user = st.selectbox("选择用户", usernames)
            with st.form("change_password_form"):
                new_password = st.text_input("新密码", type="password")
                confirm_password = st.text_input("确认新密码", type="password")
                
                if st.form_submit_button("修改"):
                    if new_password != confirm_password:
                        st.error("两次输入的密码不一致")
                    else:
                        db = get_db()
                        if db_utils.update_password(db, selected_user, new_password):
                            st.success("密码已更新")
                            st.rerun()
                        else:
                            st.error("更新失败")
        else:
            st.info("没有用户可修改")
        
        # 删除用户
        st.subheader("删除用户")
        if users:
            del_username = st.selectbox("选择要删除的用户", usernames, key="del_user_select")
            if st.button("删除用户"):
                db = get_db()
                if db_utils.delete_user(db, del_username):
                    st.success(f"用户 {del_username} 已删除")
                    st.rerun()
                else:
                    st.error("删除失败")
        else:
            st.info("没有用户可删除")

    # 值班人员管理 - 增强功能
    with st.expander("值班人员管理"):
        st.subheader("值班人员名单")
        db = get_db()
        duty_personnel = db_utils.get_all_duty_personnel(db)
        
        # 添加新值班人员
        new_person = st.text_input("添加值班人员姓名", key="new_person")
        if st.button("添加") and new_person:
            if db_utils.add_duty_person(db, new_person):
                st.success(f"已添加值班人员: {new_person}")
                st.rerun()
        
        # 显示当前值班人员并支持编辑/删除
        if duty_personnel:
            st.write("当前值班人员名单:")
            for person in duty_personnel:
                cols = st.columns([3, 1, 1])
                cols[0].write(person)
                
                # 编辑功能
                with cols[1].form(f"edit_{person}"):
                    new_name = st.text_input("新姓名", value=person, key=f"edit_name_{person}")
                    if st.form_submit_button("更新"):
                        if db_utils.update_duty_person(db, person, new_name):
                            st.success(f"已更新: {person} → {new_name}")
                            st.rerun()
                
                # 删除功能
                if cols[2].button("删除", key=f"del_{person}"):
                    if db_utils.delete_duty_person(db, person):
                        st.success(f"已删除: {person}")
                        st.rerun()
        else:
            st.warning("暂无值班人员，请先添加")

# 主工作记录页面优化布局
with tab_main:
    # 值班人员显示优化为卡片式布局
    st.markdown("### 📅 今日值班人员")
    db = get_db()
    today_duty = db_utils.get_today_duty_rotation(db)
    
    if today_duty:
        st.markdown(f"""
        <div class="card">
            <div style="display: flex; align-items: center; gap: 1rem;">
                <div style="font-size: 2rem;">👤</div>
                <div>
                    <h3 style="margin: 0; font-size: 1.2rem; color: #4b5563;">当前值班人员</h3>
                    <p style="margin: 0.25rem 0 0 0; font-size: 1.5rem; font-weight: 600; color: #1f2937;">{today_duty[0]}</p>
                </div>
            </div>
        </div>
        """, unsafe_allow_html=True)
        
        # 修改表单优化
        with st.expander("🔧 修改值班人员"):
            with st.form("edit_duty_form"):
                new_duty = st.selectbox(
                    "选择值班人员",
                    options=db_utils.get_all_duty_personnel(db),
                    index=db_utils.get_all_duty_personnel(db).index(today_duty[0]) if today_duty[0] in db_utils.get_all_duty_personnel(db) else 0,
                    key="duty_select"
                )
                
                if st.form_submit_button("保存修改"):
                    db_utils.save_today_duty(db, [new_duty])
                    st.success("今日值班人员已更新!")
                    st.rerun()
    else:
        st.warning("请先添加值班人员")

    # 工作记录管理优化为卡片布局
    st.markdown("### 📝 工作记录管理")
    tab1, tab2, tab3, tab4 = st.tabs(["➕ 添加记录", "🔍 查看/编辑记录", "📈 数据统计", "📋 待办事项"])

    with tab1:
        with st.form("add_record_form"):
            db = get_db()
            today_duty = db_utils.get_today_duty_rotation(db)
            
            recorder = st.text_input("记录人姓名")
            work_type = st.text_input("工作类型")
            work_content = st.text_area("工作内容")
            start_date = st.date_input("开始日期", value=date.today())
            end_date = st.date_input("结束日期", value=date.today())
            
            if st.form_submit_button("添加记录"):
                if recorder and work_type and work_content and start_date <= end_date:
                    db = get_db()
                    db_utils.create_record(db, recorder, work_type, work_content, start_date, end_date)
                    st.success("记录添加成功!")
                else:
                    st.error("请填写完整信息且结束日期不能早于开始日期")

    with tab2:
        db = get_db()
        records = db_utils.get_records(db)
        
        if records:
            # 分页控件优化
            col1, col2 = st.columns([3, 1])
            with col2:
                page_size = 10  # 每页显示记录数
                total_pages = (len(records) + page_size - 1) // page_size  
                page = st.number_input("页码", min_value=1, max_value=total_pages, value=1, key="record_page")
            
            # 表格容器优化
            df = pd.DataFrame([{
                "记录人": r.recorder,
                "工作类型": r.work_type,
                "工作内容": r.work_content,
                "开始日期": r.start_date,
                "结束日期": r.end_date,
                "是否完成": "是" if r.is_completed else "否"
            } for r in records])
            
            st.markdown('<div class="card">', unsafe_allow_html=True)
            st.dataframe(df, use_container_width=True, hide_index=True)
            st.markdown('</div>', unsafe_allow_html=True)
            
            # 编辑和删除区域优化
            st.markdown("#### ✏️ 编辑记录")
            record_options = {f"ID: {r.id} | {r.work_type} | {r.start_date}": r.id for r in records}
            selected_record_label = st.selectbox(
                "选择要编辑的记录",
                options=list(record_options.keys()),
                key="edit_record_select"
            )
            record_id = record_options[selected_record_label] if selected_record_label else None
            
            if record_id:
                record = next((r for r in records if r.id == record_id), None)
                if record:
                    with st.form("edit_form"):
                        new_recorder = st.text_input("记录人", value=record.recorder)
                        new_work_type = st.text_input("工作类型", value=record.work_type)
                        new_work_content = st.text_area("工作内容", value=record.work_content)
                        new_start = st.date_input("开始日期", value=record.start_date)
                        new_end = st.date_input("结束日期", value=record.end_date)
                        is_completed = st.checkbox("已完成", value=bool(record.is_completed))
                        
                        if st.form_submit_button("更新记录"):
                            if new_start <= new_end:
                                db = get_db()
                                db_utils.update_record(
                                    db, 
                                    record_id,
                                    recorder=new_recorder,
                                    work_type=new_work_type,
                                    work_content=new_work_content,
                                    start_date=new_start,
                                    end_date=new_end,
                                    is_completed=1 if is_completed else 0
                                )
                                st.success("记录更新成功!")
                                st.rerun()
                            else:
                                st.error("结束日期不能早于开始日期")
            else:
                st.warning("请选择一条记录进行编辑")

            # 删除记录优化
            st.subheader("🗑️ 删除记录")
            del_record_label = st.selectbox(
                "选择要删除的记录",
                options=list(record_options.keys()),
                key="delete_record_select"
            )
            del_id = record_options[del_record_label] if del_record_label else None
            
            if st.button("删除记录", key="delete_record_btn") and del_id:
                db = get_db()
                if db_utils.delete_record(db, del_id):
                    st.success("记录已删除!")
                    st.rerun()
                else:
                    st.error("删除失败，请检查记录状态")
        else:
            st.info("暂无工作记录")

    with tab3:
        db = get_db()
        records = db_utils.get_records(db)
        
        if records:
            # 使用Plotly生成交互式图表
            st.markdown("#### 📊 工作类型分布")
            work_types = [r.work_type for r in records]
            type_counts = pd.Series(work_types).value_counts()
            
            fig1 = px.pie(
                values=type_counts.values, 
                names=type_counts.index,
                color_discrete_sequence=px.colors.qualitative.Pastel
            )
            fig1.update_traces(textposition='inside', textinfo='percent+label')
            st.plotly_chart(fig1, use_container_width=True)
            
            st.markdown("#### 📈 记录人工作统计")
            recorders = [r.recorder for r in records]
            recorder_counts = pd.Series(recorders).value_counts()
            
            fig2 = px.bar(
                x=recorder_counts.index, 
                y=recorder_counts.values,
                color_discrete_sequence=['#636efa']
            )
            fig2.update_layout(
                xaxis_title="记录人",
                yaxis_title="记录数量",
                hovermode="x unified"
            )
            st.plotly_chart(fig2, use_container_width=True)
            
            st.markdown("#### 📅 时间分布趋势")
            df = pd.DataFrame([{
                "date": r.start_date,
                "count": 1
            } for r in records])
            
            if not df.empty:
                # 数据预处理 - 新增完整时间序列支持
                df['date'] = pd.to_datetime(df['date'])
                
                # 使用ISO周标准（周一作为周起始）
                weekly = df.set_index('date').resample('W-MON').count()
                
                # 创建完整日期范围索引并补全数据
                full_date_range = pd.date_range(
                    start=weekly.index.min() if not weekly.empty else df['date'].min(),
                    end=weekly.index.max() if not weekly.empty else df['date'].max(),
                    freq='W-MON'
                )
                weekly = weekly.reindex(full_date_range, fill_value=0).rename_axis('week_end')
                
                # 创建带滚动选择器的交互图表
                fig3 = px.line(
                    weekly.reset_index(),
                    x='week_end',  # 使用明确的周结束日期字段
                    y='count',
                    title='每周工作记录数量',
                    markers=True,  # 替换mode参数为markers参数
                    color_discrete_sequence=['#00cc96']
                )
                
                # 视觉优化配置
                fig3.update_layout(
                    xaxis_title="周结束日期",
                    yaxis_title="记录数量",
                    hovermode="x unified",
                    xaxis_rangeslider_visible=True,  # 添加滚动时间选择器
                    showlegend=False  # 移除冗余图例
                )
                fig3.update_xaxes(tickformat="%Y-%m-%d")
                
                st.plotly_chart(fig3, use_container_width=True)

    with tab4:
        # 待办事项优化
        st.markdown('<div class="card">', unsafe_allow_html=True)
        db = get_db()
        
        # 获取所有未完成的工作
        uncompleted_records = db_utils.get_uncompleted_records(db, date.today())
        if uncompleted_records:
            for record in uncompleted_records:
                with st.container(border=True):
                    cols = st.columns([4, 1])
                    cols[0].markdown(f"""
                    **工作类型**: {record.work_type}  
                    **内容**: {record.work_content}  
                    **截止日期**: {record.end_date}
                    """)
                    
                    if cols[1].button("标记完成", key=f"complete_{record.id}"):
                        db_utils.update_record(db, record.id, is_completed=1)
                        st.rerun()
        else:
            st.success("当前没有待办工作")

# Excel导出功能优化
st.markdown("### 📦 导出工作记录")
col1, col2 = st.columns(2)
with col1:
    export_start = st.date_input("起始日期", value=date.today() - timedelta(days=30))
with col2:
    export_end = st.date_input("结束日期", value=date.today())

if st.button("📥 导出为Excel", use_container_width=True):
    db = get_db()
    df = db_utils.export_to_excel(db, export_start, export_end)
    
    if not df.empty:
        # 创建Excel文件
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='工作记录')
            
            # 获取工作表
            worksheet = writer.sheets['工作记录']
            
            # 设置表头样式
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # 应用表头样式
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment
            
            # 自动调整列宽
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_letter].width = min(adjusted_width, 50)  # 限制最大宽度
            
            # 冻结首行
            worksheet.freeze_panes = 'A2'

        output.seek(0)
        
        # 提供下载
        st.download_button(
            label="下载Excel文件",
            data=output,
            file_name=f"work_records_{export_start}_{export_end}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.warning("所选时间段内没有记录")

# 侧边栏提醒部分 - 移动到主界面之外
with st.sidebar:
    # 添加: 强化提醒条件判断
    if 'show_pending_records' in st.session_state and st.session_state.show_pending_records:
        st.markdown("### ⚠️ 待处理工作提醒")
        
        # 获取最新未完成记录（防止数据陈旧）
        db = get_db()
        current_pending = db_utils.get_uncompleted_records(db)
        
        if current_pending:
            for record in current_pending:
                st.markdown(f"""
                <div class="reminder-card">
                    <div style="display: flex; justify-content: space-between; align-items: center;">
                        <div>
                            <strong class="icon-animation">📌 {record.work_type}</strong><br>
                            <small>截止: {record.end_date}</small>
                        </div>
                        <div style="font-size: 1.5rem; color: #ea580c;">❗</div>
                    </div>
                </div>
                """, unsafe_allow_html=True)
                    
                if st.button(f"✅ 标记为已完成", key=f"sidebar_complete_{record.id}", use_container_width=True):
                        db = get_db()
                        db_utils.update_record(db, record.id, is_completed=1)
                        st.session_state.pending_records = [
                            r for r in st.session_state.pending_records if r.id != record.id
                        ]
                        st.toast(f"记录 {record.id} 已标记为完成", icon='✅')
                        st.rerun()
        else:
            st.info("暂无待处理工作")




















