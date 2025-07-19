from datetime import date, timedelta
from io import BytesIO

import pandas as pd
import plotly.express as px
import streamlit as st
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import db_utils

# 在全局样式部分添加备份按钮样式
st.markdown("""
<style>
    /* 备份按钮样式 */
    .backup-btn {
        background: linear-gradient(45deg, #4CAF50, #8BC34A);
        color: white;
        border-radius: 10px;
        padding: 8px 15px;
        font-weight: bold;
    }
    .backup-btn:hover {
        transform: scale(1.05);
    }
</style>
""", unsafe_allow_html=True)

def show_work_record_page():
    """展示工作记录管理页面"""
    st.markdown("### 📝 工作记录管理")
    
    # 功能卡片导航
    cols = st.columns(4)
    with cols[0]:
        if st.button("➕ 添加记录", use_container_width=True, key="add_record_btn"):
            st.session_state.current_work_record_view = "add"
    with cols[1]:
        if st.button("🔍 查看/编辑", use_container_width=True, key="edit_record_btn"):
            st.session_state.current_work_record_view = "edit"
    with cols[2]:
        if st.button("📈 数据统计", use_container_width=True, key="stats_btn"):
            st.session_state.current_work_record_view = "stats"
    with cols[3]:
        if st.button("📋 待办事项", use_container_width=True, key="todo_btn"):
            st.session_state.current_work_record_view = "todo"

    # 根据选择显示对应功能
    if 'current_work_record_view' not in st.session_state:
        st.session_state.current_work_record_view = "add"
    
    if st.session_state.current_work_record_view == "add":
        show_add_record_form()
    elif st.session_state.current_work_record_view == "edit":
        show_edit_records()
    elif st.session_state.current_work_record_view == "stats":
        show_statistics()
    elif st.session_state.current_work_record_view == "todo":
        show_todo_list()

def show_add_record_form():
    """展示添加记录表单"""
    with st.form("add_record_form"):
        db = next(db_utils.get_db_session())
        today_duty = db_utils.get_today_duty_rotation(db)
        
        recorder = st.text_input("记录人姓名")
        work_type = st.text_input("工作类型")
        work_content = st.text_area("工作内容")
        start_date = st.date_input("开始日期", value=date.today())
        end_date = st.date_input("结束日期", value=date.today())
        
        if st.form_submit_button("添加记录"):
            if recorder and work_type and work_content and start_date <= end_date:
                db = next(db_utils.get_db_session())
                db_utils.create_record(db, recorder, work_type, work_content, start_date, end_date)
                st.success("记录添加成功!")
            else:
                st.error("请填写完整信息且结束日期不能早于开始日期")

def show_edit_records():
    """展示编辑记录界面"""
    db = next(db_utils.get_db_session())
    records = db_utils.get_records(db)
    
    if records:
        # 分页控件
        col1, col2 = st.columns([3, 1])
        with col2:
            page_size = 10
            total_pages = (len(records) + page_size - 1) // page_size  
            page = st.number_input("页码", min_value=1, max_value=total_pages, value=1, key="record_page")
        
        # 表格展示
        df = pd.DataFrame([{
            "记录人": r.recorder,
            "工作类型": r.work_type,
            "工作内容": r.work_content,
            "开始日期": r.start_date,
            "结束日期": r.end_date,
            "是否完成": "是" if r.is_completed else "否"
        } for r in records])
        
        st.dataframe(df, use_container_width=True, hide_index=True)
        
        # 编辑和删除区域
        st.markdown("#### ✏️ 编辑记录")
        record_options = {f"ID: {r.id} | 工作类型: {r.work_type} | 工作内容: {r.work_content} |开始日期: {r.start_date}": r.id for r in records}
        selected_record_label = st.selectbox(
            "选择要编辑的记录",
            options=list(record_options.keys()),
            key="edit_record_select"
        )
        record_id = record_options[selected_record_label] if selected_record_label else None
        
        if record_id:
            record = next((r for r in records if r.id == record_id), None)
            if record:
                with st.form("edit_form"):
                    new_recorder = st.text_input("记录人", value=record.recorder)
                    new_work_type = st.text_input("工作类型", value=record.work_type)
                    new_work_content = st.text_area("工作内容", value=record.work_content)
                    new_start = st.date_input("开始日期", value=record.start_date)
                    new_end = st.date_input("结束日期", value=record.end_date)
                    is_completed = st.checkbox("已完成", value=bool(record.is_completed))
                    
                    if st.form_submit_button("更新记录"):
                        if new_start <= new_end:
                            db = next(db_utils.get_db_session())
                            db_utils.update_record(
                                db, 
                                record_id,
                                recorder=new_recorder,
                                work_type=new_work_type,
                                work_content=new_work_content,
                                start_date=new_start,
                                end_date=new_end,
                                is_completed=1 if is_completed else 0
                            )
                            st.success("记录更新成功!")
                            st.rerun()
                        else:
                            st.error("结束日期不能早于开始日期")
        else:
            st.warning("请选择一条记录进行编辑")

        # 删除记录
        st.subheader("🗑️ 删除记录")
        del_record_label = st.selectbox(
            "选择要删除的记录",
            options=list(record_options.keys()),
            key="delete_record_select"
        )
        del_id = record_options[del_record_label] if del_record_label else None
        
        if st.button("删除记录", key="delete_record_btn") and del_id:
            db = next(db_utils.get_db_session())
            if db_utils.delete_record(db, del_id):
                st.success("记录已删除!")
                st.rerun()
            else:
                st.error("删除失败，请检查记录状态")
    else:
        st.info("暂无工作记录")

def show_statistics():
    """展示统计数据图表"""
    db = next(db_utils.get_db_session())
    records = db_utils.get_records(db)
    
    if records:
        # 工作类型分布
        st.markdown("#### 📊 工作类型分布")
        work_types = [r.work_type for r in records]
        type_counts = pd.Series(work_types).value_counts()
        
        fig1 = px.pie(
            values=type_counts.values, 
            names=type_counts.index,
            color_discrete_sequence=px.colors.qualitative.Pastel
        )
        fig1.update_traces(textposition='inside', textinfo='percent+label')
        st.plotly_chart(fig1, use_container_width=True)
        
        # 记录人工作统计
        st.markdown("#### 📈 记录人工作统计")
        recorders = [r.recorder for r in records]
        recorder_counts = pd.Series(recorders).value_counts()
        
        fig2 = px.bar(
            x=recorder_counts.index, 
            y=recorder_counts.values,
            color_discrete_sequence=['#636efa']
        )
        fig2.update_layout(
            xaxis_title="记录人",
            yaxis_title="记录数量",
            hovermode="x unified"
        )
        st.plotly_chart(fig2, use_container_width=True)
        
        # 时间分布趋势
        st.markdown("#### 📅 时间分布趋势")
        df = pd.DataFrame([{
            "date": r.start_date,
            "count": 1
        } for r in records])
        
        if not df.empty:
            df['date'] = pd.to_datetime(df['date'])
            weekly = df.set_index('date').resample('W-MON').count()
            full_date_range = pd.date_range(
                start=weekly.index.min() if not weekly.empty else df['date'].min(),
                end=weekly.index.max() if not weekly.empty else df['date'].max(),
                freq='W-MON'
            )
            weekly = weekly.reindex(full_date_range, fill_value=0).rename_axis('week_end')
            
            fig3 = px.line(
                weekly.reset_index(),
                x='week_end',
                y='count',
                title='每周工作记录数量',
                markers=True,
                color_discrete_sequence=['#00cc96']
            )
            
            fig3.update_layout(
                xaxis_title="周结束日期",
                yaxis_title="记录数量",
                hovermode="x unified",
                xaxis_rangeslider_visible=True,
                showlegend=False
            )
            fig3.update_xaxes(tickformat="%Y-%m-%d")
            
            st.plotly_chart(fig3, use_container_width=True)

def show_todo_list():
    """展示待办事项"""
    db = next(db_utils.get_db_session())
    uncompleted_records = db_utils.get_uncompleted_records(db, date.today())
    if uncompleted_records:
        for record in uncompleted_records:
            with st.container(border=True):
                cols = st.columns([4, 1])
                cols[0].markdown(f"""
                **记录人**: {record.recorder}\n
                **工作类型**: {record.work_type}\n
                **内容**: {record.work_content}\n 
                **截止日期**: {record.end_date}
                """)
                
                if cols[1].button("标记完成", key=f"complete_{record.id}"):
                    db_utils.update_record(db, record.id, is_completed=1)
                    st.rerun()
    else:
        st.success("当前没有待办工作")

def show_export_section():
    """展示导出功能"""
    st.markdown("### 📦 导出工作记录")
    col1, col2 = st.columns(2)
    with col1:
        export_start = st.date_input("起始日期", value=date.today() - timedelta(days=30))
    with col2:
        export_end = st.date_input("结束日期", value=date.today())

    if st.button("📥 导出为Excel", use_container_width=True):
        db = next(db_utils.get_db_session())
        df = db_utils.export_to_excel(db, export_start, export_end)
        
        if not df.empty:
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='工作记录')
                worksheet = writer.sheets['工作记录']
                
                header_font = Font(bold=True, color='FFFFFF')
                header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
                alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = alignment
                
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column_letter].width = min(adjusted_width, 50)
                
                worksheet.freeze_panes = 'A2'

            output.seek(0)
            
            st.download_button(
                label="下载Excel文件",
                data=output,
                file_name=f"work_records_{export_start}_{export_end}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.warning("所选时间段内没有记录")